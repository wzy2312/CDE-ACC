import json
import os
import sys
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas


FONT_PATH = "/System/Library/Fonts/Supplemental/Arial Unicode.ttf"
FONT_NAME = "ScheduleArialUnicode"
PAGE_WIDTH, PAGE_HEIGHT = A4
MARGIN_X = 46
TOP_Y = PAGE_HEIGHT - 54
BOTTOM_Y = 48
LINE_HEIGHT = 17
HEADER_FILL = PatternFill("solid", fgColor="E8EEF7")
TITLE_FILL = PatternFill("solid", fgColor="DDE8F5")
WARN_FILL = PatternFill("solid", fgColor="FCE8E6")
SUCCESS_FILL = PatternFill("solid", fgColor="E8F6EF")
TEXT_WRAP = Alignment(vertical="top", wrap_text=True)


def main():
    if len(sys.argv) < 5:
        raise SystemExit("usage: export_construction_schedule.py <store_json> <schedule_id> <kind> <output> [week_start] [week_end]")

    store_path, schedule_id, kind, output_path = sys.argv[1:5]
    week_start = safe_text(sys.argv[5] if len(sys.argv) > 5 else "")
    week_end = safe_text(sys.argv[6] if len(sys.argv) > 6 else "")
    with open(store_path, "r", encoding="utf-8") as handle:
        store = json.load(handle)

    schedule = find_by_id(store.get("constructionScheduleVersions", []), schedule_id)
    if not schedule:
        raise SystemExit("construction schedule not found in store")

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    if kind == "workbook":
        export_workbook(output_path, store, schedule)
    elif kind == "weekly_pdf":
        export_weekly_pdf(output_path, store, schedule, week_start, week_end)
    elif kind == "gantt_pdf":
        export_gantt_pdf(output_path, store, schedule)
    else:
        raise SystemExit(f"unsupported construction schedule export kind: {kind}")


def safe_text(value, fallback=""):
    if value is None:
        return fallback
    text = str(value).strip()
    return text if text else fallback


def number(value, fallback=0):
    try:
        return float(value)
    except Exception:
        return fallback


def find_by_id(items, item_id):
    return next((item for item in items if safe_text(item.get("id")) == safe_text(item_id)), None)


def schedule_records(store, schedule, key):
    return [item for item in store.get(key, []) if safe_text(item.get("scheduleId")) == safe_text(schedule.get("id"))]


def document_for_schedule(store, schedule):
    return find_by_id(store.get("documents", []), schedule.get("documentId")) or {}


def parse_date(value):
    raw = safe_text(value)
    if not raw:
        return None
    if len(raw) >= 10:
        try:
            return datetime.strptime(raw[:10], "%Y-%m-%d")
        except Exception:
            pass
    for fmt in ("%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(raw, fmt)
        except Exception:
            pass
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00")).replace(tzinfo=None)
    except Exception:
        return None


def format_date(value):
    dt = parse_date(value)
    return dt.strftime("%Y-%m-%d") if dt else safe_text(value)


def format_datetime(value):
    dt = parse_date(value)
    return dt.strftime("%Y-%m-%d %H:%M") if dt else safe_text(value)


def status_label(value):
    return {
        "Not Started": "未开始",
        "In Progress": "进行中",
        "Completed": "已完成",
        "future": "未来计划",
        "not_started": "未开始",
        "in_progress": "进行中",
        "completed": "已完成",
        "early": "提前完成",
        "delayed": "滞后",
    }.get(safe_text(value), safe_text(value, "未知"))


def activity_status(activity, date_value):
    view_date = parse_date(date_value) or datetime.utcnow()
    planned_start = parse_date(activity.get("plannedStart"))
    planned_finish = parse_date(activity.get("plannedFinish"))
    actual_start = parse_date(activity.get("actualStart"))
    actual_finish = parse_date(activity.get("actualFinish"))
    pct = number(activity.get("percentComplete"))
    if actual_finish or pct >= 100 or safe_text(activity.get("status")) == "Completed":
        if actual_finish and planned_finish and actual_finish < planned_finish:
            return "early"
        return "completed"
    if planned_start and view_date < planned_start:
        return "future"
    if planned_finish and view_date > planned_finish:
        return "delayed"
    if actual_start or pct > 0 or safe_text(activity.get("status")) == "In Progress":
        return "in_progress"
    if planned_start and view_date >= planned_start:
        return "delayed"
    return "not_started"


def write_meta(sheet, document, schedule):
    sheet["A1"] = "海水淡化项目施工进度计划"
    sheet["A1"].font = Font(bold=True, size=16, color="172033")
    sheet["A1"].fill = TITLE_FILL
    sheet.merge_cells("A1:H1")
    rows = [
        ("模型名称", safe_text(document.get("name"), "未命名模型"), "计划版本", safe_text(schedule.get("name"))),
        ("项目名称", safe_text(schedule.get("projectName")), "P6 项目", safe_text(schedule.get("projectCode") or schedule.get("p6ProjectId"))),
        ("数据日期", safe_text(schedule.get("dataDate")), "计划范围", f"{safe_text(schedule.get('plannedStart'))} → {safe_text(schedule.get('plannedFinish'))}"),
        ("导入人", safe_text(schedule.get("importedBy")), "导入时间", format_datetime(schedule.get("importedAt"))),
    ]
    for row_index, row in enumerate(rows, start=2):
        for col_index, value in enumerate(row, start=1):
            cell = sheet.cell(row_index, col_index, value)
            cell.alignment = TEXT_WRAP
            if col_index in (1, 3):
                cell.font = Font(bold=True)
                cell.fill = HEADER_FILL


def write_header(sheet, row, headers):
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row, col, header)
        cell.font = Font(bold=True, color="172033")
        cell.fill = HEADER_FILL
        cell.alignment = TEXT_WRAP


def write_row(sheet, row, values, fill=None):
    for col, value in enumerate(values, start=1):
        cell = sheet.cell(row, col, value)
        cell.alignment = TEXT_WRAP
        if fill:
            cell.fill = fill


def fit_columns(sheet, widths):
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def export_workbook(output_path, store, schedule):
    document = document_for_schedule(store, schedule)
    activities = schedule_records(store, schedule, "constructionScheduleActivities")
    mappings = schedule_records(store, schedule, "constructionScheduleMappings")
    reports = schedule_records(store, schedule, "constructionProgressReports")
    alerts = schedule_records(store, schedule, "constructionScheduleAlerts")

    workbook = Workbook()
    activity_sheet = workbook.active
    activity_sheet.title = "Activity计划与实际"
    write_meta(activity_sheet, document, schedule)
    activity_headers = ["Activity ID", "名称", "WBS", "类型", "计划开始", "计划完成", "实际开始", "实际完成", "完成率", "状态", "总浮时(h)", "关联构件数"]
    write_header(activity_sheet, 7, activity_headers)
    mapping_count = {}
    for mapping in mappings:
        mapping_count[mapping.get("activityId")] = mapping_count.get(mapping.get("activityId"), 0) + 1
    for row_index, activity in enumerate(sorted(activities, key=lambda item: safe_text(item.get("activityId"))), start=8):
        fill = WARN_FILL if activity_status(activity, schedule.get("dataDate")) == "delayed" else None
        write_row(activity_sheet, row_index, [
            activity.get("activityId"),
            activity.get("name"),
            activity.get("wbsId"),
            activity.get("type"),
            activity.get("plannedStart"),
            activity.get("plannedFinish"),
            activity.get("actualStart"),
            activity.get("actualFinish"),
            activity.get("percentComplete"),
            status_label(activity.get("status")),
            activity.get("totalFloatHours"),
            mapping_count.get(activity.get("activityId"), 0),
        ], fill)
    activity_sheet.freeze_panes = "A8"
    fit_columns(activity_sheet, [14, 34, 16, 16, 14, 14, 14, 14, 10, 12, 12, 12])

    mapping_sheet = workbook.create_sheet("构件-工序映射")
    write_header(mapping_sheet, 1, ["Activity ID", "dbId", "UniqueId/Handle", "构件名称", "对象类型", "系统/区域", "专业", "方式", "置信度", "原因"])
    for row_index, mapping in enumerate(mappings, start=2):
        write_row(mapping_sheet, row_index, [
            mapping.get("activityId"),
            mapping.get("dbId"),
            mapping.get("uniqueId"),
            mapping.get("elementName"),
            mapping.get("elementType"),
            mapping.get("floor"),
            mapping.get("discipline"),
            mapping.get("matchMethod"),
            mapping.get("confidence"),
            mapping.get("reason"),
        ])
    fit_columns(mapping_sheet, [14, 10, 28, 26, 18, 18, 14, 16, 12, 36])

    report_sheet = workbook.create_sheet("现场上报")
    write_header(report_sheet, 1, ["Activity ID", "上报人", "上报时间", "实际开始", "实际完成", "完成率", "说明"])
    for row_index, report in enumerate(reports, start=2):
        write_row(report_sheet, row_index, [
            report.get("activityId"),
            report.get("reportedBy"),
            format_datetime(report.get("reportedAt")),
            report.get("actualStart"),
            report.get("actualFinish"),
            report.get("percentComplete"),
            report.get("note"),
        ])
    fit_columns(report_sheet, [14, 18, 18, 14, 14, 10, 52])

    alert_sheet = workbook.create_sheet("进度预警")
    write_header(alert_sheet, 1, ["Activity ID", "类型", "严重度", "计划完成", "完成率", "滞后天数", "状态", "Issue", "说明"])
    for row_index, alert in enumerate(alerts, start=2):
        fill = WARN_FILL if safe_text(alert.get("status")) == "open" else SUCCESS_FILL
        write_row(alert_sheet, row_index, [
            alert.get("activityCode") or alert.get("activityId"),
            alert.get("type"),
            alert.get("severity"),
            alert.get("plannedFinish"),
            alert.get("percentComplete"),
            alert.get("delayDays"),
            alert.get("status"),
            alert.get("issueId"),
            alert.get("message"),
        ], fill)
    fit_columns(alert_sheet, [14, 16, 12, 14, 10, 10, 12, 24, 60])

    summary_sheet = workbook.create_sheet("导入变更摘要")
    write_header(summary_sheet, 1, ["指标", "数量"])
    import_summary = schedule.get("importSummary") or {}
    rows = [
        ("新增 Activity", import_summary.get("added", 0)),
        ("删除 Activity", import_summary.get("deleted", 0)),
        ("日期调整", import_summary.get("dateChanged", 0)),
        ("状态变化", import_summary.get("statusChanged", 0)),
        ("进度变化", import_summary.get("progressChanged", 0)),
        ("未变化", import_summary.get("unchanged", 0)),
    ]
    for row_index, row in enumerate(rows, start=2):
        write_row(summary_sheet, row_index, row)
    write_header(summary_sheet, 10, ["Activity ID", "名称", "变更类型", "变更明细"])
    for row_index, item in enumerate(import_summary.get("changedActivities", [])[:500], start=11):
        changes = "；".join(f"{change.get('field')}: {change.get('valueBefore')} -> {change.get('valueAfter')}" for change in item.get("changes", []) or [])
        write_row(summary_sheet, row_index, [item.get("activityId"), item.get("name"), item.get("type"), changes])
    fit_columns(summary_sheet, [18, 34, 16, 80])

    workbook.save(output_path)


def register_font():
    try:
        pdfmetrics.registerFont(TTFont(FONT_NAME, FONT_PATH))
        return FONT_NAME
    except Exception:
        return "Helvetica"


def draw_text_lines(pdf, cursor_y, title, lines):
    if cursor_y < BOTTOM_Y + 80:
        pdf.showPage()
        pdf.setFont(register_font(), 10)
        cursor_y = TOP_Y
    pdf.setFont(register_font(), 12)
    pdf.drawString(MARGIN_X, cursor_y, title)
    cursor_y -= LINE_HEIGHT + 4
    pdf.setFont(register_font(), 9)
    for line in lines:
        for chunk in wrap_text(safe_text(line, "—"), 86):
            if cursor_y < BOTTOM_Y:
                pdf.showPage()
                pdf.setFont(register_font(), 9)
                cursor_y = TOP_Y
            pdf.drawString(MARGIN_X, cursor_y, chunk)
            cursor_y -= LINE_HEIGHT
    return cursor_y - 8


def wrap_text(text, limit):
    return [text[index:index + limit] for index in range(0, len(text), limit)] or [""]


def weekly_stats(activities, week_start, week_end):
    start_dt = parse_date(week_start) or datetime.utcnow()
    end_dt = parse_date(week_end) or start_dt + timedelta(days=6)
    total = max(1, len([item for item in activities if safe_text(item.get("type")) != "Milestone"]))
    planned_done = [item for item in activities if parse_date(item.get("plannedFinish")) and parse_date(item.get("plannedFinish")) <= end_dt and safe_text(item.get("type")) != "Milestone"]
    actual_percent = sum(number(item.get("percentComplete")) for item in activities if safe_text(item.get("type")) != "Milestone") / total
    planned_percent = round(len(planned_done) / total * 100, 1)
    unfinished = [item for item in planned_done if number(item.get("percentComplete")) < 100 and not safe_text(item.get("actualFinish"))]
    next_plan = [item for item in activities if parse_date(item.get("plannedStart")) and end_dt < parse_date(item.get("plannedStart")) <= end_dt + timedelta(days=30)]
    return planned_percent, round(actual_percent, 1), unfinished, next_plan


def export_weekly_pdf(output_path, store, schedule, week_start, week_end):
    font = register_font()
    activities = schedule_records(store, schedule, "constructionScheduleActivities")
    document = document_for_schedule(store, schedule)
    start = week_start or safe_text(schedule.get("dataDate"))
    end = week_end or (parse_date(start) + timedelta(days=6)).strftime("%Y-%m-%d") if parse_date(start) else start
    planned, actual, unfinished, next_plan = weekly_stats(activities, start, end)
    pdf = canvas.Canvas(output_path, pagesize=A4)
    pdf.setTitle(f"Construction Schedule Weekly Report - {safe_text(schedule.get('name'))}")
    pdf.setFont(font, 14)
    pdf.drawString(MARGIN_X, TOP_Y, "海水淡化项目进度周报")
    cursor = TOP_Y - 32
    cursor = draw_text_lines(pdf, cursor, "基本信息", [
        f"模型：{safe_text(document.get('name'), '未命名模型')}",
        f"计划版本：{safe_text(schedule.get('name'))}",
        f"数据日期：{safe_text(schedule.get('dataDate'))}",
        f"周报周期：{start} 至 {end}",
    ])
    cursor = draw_text_lines(pdf, cursor, "总体进度", [
        f"计划进度：{planned}%",
        f"实际进度：{actual}%",
        f"进度偏差：{round(actual - planned, 1)}%",
    ])
    cursor = draw_text_lines(pdf, cursor, "本周计划未完成", [
        f"{item.get('activityId')} {item.get('name')}（{item.get('percentComplete')}%）"
        for item in unfinished[:16]
    ] or ["无"])
    draw_text_lines(pdf, cursor, "下阶段计划", [
        f"{item.get('activityId')} {item.get('name')}（{item.get('plannedStart')} → {item.get('plannedFinish')}）"
        for item in next_plan[:16]
    ] or ["无"])
    pdf.save()


def export_gantt_pdf(output_path, store, schedule):
    font = register_font()
    activities = sorted(schedule_records(store, schedule, "constructionScheduleActivities"), key=lambda item: safe_text(item.get("plannedStart")) + safe_text(item.get("activityId")))
    pdf = canvas.Canvas(output_path, pagesize=A4)
    pdf.setTitle(f"Construction Gantt - {safe_text(schedule.get('name'))}")
    pdf.setFont(font, 14)
    pdf.drawString(MARGIN_X, TOP_Y, "施工进度甘特导出")
    cursor = TOP_Y - 28
    pdf.setFont(font, 9)
    for index, activity in enumerate(activities[:80], start=1):
      if cursor < BOTTOM_Y:
          pdf.showPage()
          pdf.setFont(font, 9)
          cursor = TOP_Y
      status = activity_status(activity, schedule.get("dataDate"))
      pdf.drawString(MARGIN_X, cursor, f"{index}. {activity.get('activityId')} {activity.get('name')}")
      pdf.drawRightString(PAGE_WIDTH - MARGIN_X, cursor, f"{activity.get('plannedStart')} -> {activity.get('plannedFinish')} · {status_label(status)} · {activity.get('percentComplete')}%")
      cursor -= LINE_HEIGHT
    pdf.save()


if __name__ == "__main__":
    main()
