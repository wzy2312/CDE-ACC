import json
import os
import sys
from collections import Counter
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from export_comment_report import (
    base_row,
    collect_annotation_rows,
    collect_native_comment_rows,
    collect_remark_rows,
    current_version_entry,
    safe_text,
    sort_key_for_row,
    write_comment_sheet,
)
from report_i18n import localize_workbook


def main():
    if len(sys.argv) != 5:
        raise SystemExit("usage: export_workflow_comment_report.py <store_json> <workflow_id> <manifest_json> <output_xlsx>")

    store_path, workflow_id, manifest_path, output_xlsx = sys.argv[1:]

    with open(store_path, "r", encoding="utf-8") as handle:
        store = json.load(handle)

    workflows = [] if isinstance(store, list) else store.get("workflows", [])
    documents = [] if isinstance(store, list) else store.get("documents", [])

    workflow = next((item for item in workflows if safe_text(item.get("id")) == safe_text(workflow_id)), None)
    if not workflow:
        raise SystemExit("workflow not found in store")

    with open(manifest_path, "r", encoding="utf-8") as handle:
        manifest = json.load(handle)

    manifest_docs = manifest.get("documents", []) if isinstance(manifest, dict) else []
    documents_by_id = {safe_text(item.get("id")): item for item in documents}
    rows = []
    native_support = []
    exported_docs = []

    for manifest_doc in manifest_docs:
      doc_id = safe_text(manifest_doc.get("docId"))
      input_file = safe_text(manifest_doc.get("inputFile"))
      document = documents_by_id.get(doc_id)
      if not document or not input_file:
          continue
      exported_docs.append(document)
      version_entry = current_version_entry(document)
      ext = os.path.splitext(version_entry.get("name") or document.get("name") or input_file)[1].lower()
      native_rows, support = collect_native_comment_rows(input_file, ext, document, version_entry)
      rows.extend(collect_remark_rows(document, version_entry))
      rows.extend(collect_annotation_rows(document, safe_text(document.get("currentVersionId")), version_entry))
      rows.extend(native_rows)
      native_support.append((safe_text(document.get("name"), doc_id), support))

    rows.extend(collect_workflow_comment_rows(workflow))
    rows.sort(key=sort_key_for_row)

    workbook = Workbook()
    comment_sheet = workbook.active
    comment_sheet.title = "评论清单"
    write_comment_sheet(comment_sheet, rows)

    summary_sheet = workbook.create_sheet("导出摘要")
    write_workflow_summary_sheet(summary_sheet, workflow, exported_docs, rows, native_support)

    localize_workbook(workbook)
    workbook.save(output_xlsx)


def collect_workflow_comment_rows(workflow):
    rows = []
    pseudo_document = {
        "name": safe_text(workflow.get("workflowName"), "未命名流程"),
    }
    pseudo_version = {
        "version": "流程级",
        "id": safe_text(workflow.get("id")),
        "name": safe_text(workflow.get("workflowName"), "未命名流程"),
    }
    for step in workflow.get("steps") or []:
        step_name = safe_text(step.get("name"), "审批节点")
        for reviewer in step.get("reviewers") or []:
            comment = safe_text(reviewer.get("comment"))
            if not comment:
                continue
            rows.append(
                base_row(
                    pseudo_document,
                    pseudo_version,
                    "流程审批意见",
                    "审批流程",
                    record_id=reviewer.get("id"),
                    status=safe_text(reviewer.get("status")),
                    actor=safe_text(reviewer.get("name"), "未知审批人"),
                    created_at=reviewer.get("actedAt"),
                    page_or_sheet=step_name,
                    anchor=safe_text(workflow.get("workflowName")),
                    title=" / ".join(
                        [
                            item
                            for item in [
                                safe_text(workflow.get("workflowName")),
                                step_name,
                            ]
                            if item
                        ]
                    ),
                    content=comment,
                    version_id=safe_text(workflow.get("id")),
                )
            )
    return rows


def write_workflow_summary_sheet(sheet, workflow, documents, rows, native_support):
    counts_by_source = Counter(row.get("评论来源") for row in rows)
    counts_by_type = Counter(row.get("记录类型") for row in rows)
    document_names = " / ".join(safe_text(item.get("name")) for item in documents[:6] if safe_text(item.get("name")))
    if len(documents) > 6:
        document_names = f"{document_names} 等 {len(documents)} 个文件"
    meta_rows = [
        ("流程名称", safe_text(workflow.get("workflowName"), "未命名流程")),
        ("导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("文件数量", str(len(documents))),
        ("文件范围", document_names or "未关联文件"),
        ("评论总数", str(len(rows))),
        ("说明", "已汇总流程下所有文件的系统批注、回复、备注、Office 原生评论，以及流程审批意见。"),
        ("", ""),
        ("按来源统计", "数量"),
    ]
    for source, count in sorted(counts_by_source.items()):
        meta_rows.append((source, str(count)))
    meta_rows.append(("", ""))
    meta_rows.append(("按类型统计", "数量"))
    for record_type, count in sorted(counts_by_type.items()):
        meta_rows.append((record_type, str(count)))
    meta_rows.append(("", ""))
    meta_rows.append(("原生评论提取情况", "结果"))
    for doc_name, support in native_support:
        meta_rows.append((doc_name, support))

    for index, (label, value) in enumerate(meta_rows, start=1):
        sheet.append([label, value])
        if label in {"按来源统计", "按类型统计", "原生评论提取情况"}:
            sheet[f"A{index}"].font = Font(bold=True, color="FFFFFF")
            sheet[f"B{index}"].font = Font(bold=True, color="FFFFFF")
            sheet[f"A{index}"].fill = PatternFill("solid", fgColor="1F5AA6")
            sheet[f"B{index}"].fill = PatternFill("solid", fgColor="1F5AA6")
        elif label:
            sheet[f"A{index}"].font = Font(bold=True)
        sheet[f"A{index}"].alignment = Alignment(vertical="top", wrap_text=True)
        sheet[f"B{index}"].alignment = Alignment(vertical="top", wrap_text=True)

    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 80


if __name__ == "__main__":
    main()
