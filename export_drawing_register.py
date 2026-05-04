import json
import os
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="E8EEF7")
TITLE_FILL = PatternFill("solid", fgColor="DDE8F5")
WARNING_FILL = PatternFill("solid", fgColor="FFF4D8")
PASS_FILL = PatternFill("solid", fgColor="EAF7EE")
TEXT_WRAP = Alignment(vertical="top", wrap_text=True)


def main():
    if len(sys.argv) != 3:
        raise SystemExit("usage: export_drawing_register.py <manifest_json> <output_xlsx>")

    manifest_path, output_xlsx = sys.argv[1:]
    with open(manifest_path, "r", encoding="utf-8") as handle:
        manifest = json.load(handle)

    register = manifest.get("register") or {}
    workbook = Workbook()
    register_sheet = workbook.active
    register_sheet.title = "图纸目录"
    write_register_sheet(register_sheet, register.get("entries") or [])

    package_sheet = workbook.create_sheet("打印包")
    write_package_sheet(package_sheet, register.get("packages") or [])

    summary_sheet = workbook.create_sheet("导出摘要")
    write_summary_sheet(summary_sheet, register)

    os.makedirs(os.path.dirname(output_xlsx), exist_ok=True)
    workbook.save(output_xlsx)


def safe_text(value, fallback=""):
    if value is None:
        return fallback
    text = str(value).strip()
    return text if text else fallback


def format_timestamp(value):
    raw = safe_text(value)
    if not raw:
        return ""
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00")).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return raw


def format_status(exists):
    return "存在" if exists else "缺失"


def package_refs_label(version):
    refs = version.get("packageRefs") if isinstance(version.get("packageRefs"), list) else []
    return "、".join(safe_text(item.get("packageNo")) for item in refs if safe_text(item.get("packageNo")))


def consistency_label(version):
    status = safe_text(version.get("consistencyStatus"))
    return {
        "pass": "通过",
        "warning": "警告",
        "confirmed": "已确认",
        "single": "—",
    }.get(status, "—")


def metadata_source_label(version):
    source = safe_text(version.get("metadataSource"))
    if source == "auto" or version.get("requiresMetadataConfirmation"):
        return "自动归集"
    if source == "declared":
        return "用户声明"
    return "未识别"


def write_register_sheet(sheet, entries):
    headers = [
        "图纸编号",
        "图纸名称",
        "专业",
        "归集来源",
        "当前版本",
        "版本日期",
        "审批状态",
        "PDF状态",
        "DWG状态",
        "格式一致性",
        "所属打印包编号",
        "备注",
    ]
    write_header(sheet, 1, headers)
    for row_index, entry in enumerate(entries, start=2):
        version = entry.get("currentVersion") or {}
        formats = version.get("formats") if isinstance(version.get("formats"), dict) else {}
        warnings = version.get("consistencyDetails") if isinstance(version.get("consistencyDetails"), list) else []
        row = [
            safe_text(entry.get("drawingNo")),
            safe_text(entry.get("name")),
            safe_text(entry.get("discipline")),
            metadata_source_label(version),
            safe_text(version.get("version")),
            format_timestamp(version.get("versionDate")),
            safe_text(version.get("approvalStatus")),
            format_status(bool(formats.get("pdf"))),
            format_status(bool(formats.get("dwg"))),
            consistency_label(version),
            package_refs_label(version),
            "；".join(safe_text(item.get("message") or f"{item.get('field')}：PDF={item.get('pdfValue')} / DWG={item.get('dwgValue')}") for item in warnings),
        ]
        write_row(sheet, row_index, row)
        if safe_text(version.get("consistencyStatus")) == "warning":
            for col_index in range(1, len(headers) + 1):
                sheet.cell(row_index, col_index).fill = WARNING_FILL
        elif safe_text(version.get("consistencyStatus")) == "pass":
            sheet.cell(row_index, 10).fill = PASS_FILL

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:L{max(2, len(entries) + 1)}"
    fit_columns(sheet, [18, 28, 14, 14, 12, 20, 14, 12, 12, 14, 24, 46])


def write_package_sheet(sheet, packages):
    headers = ["包编号", "包名称", "类型", "图纸数", "状态", "创建人", "创建时间", "签收完成", "预检查"]
    write_header(sheet, 1, headers)
    for row_index, package in enumerate(packages, start=2):
        warnings = package.get("preCheckWarnings") if isinstance(package.get("preCheckWarnings"), list) else []
        row = [
            safe_text(package.get("packageNo")),
            safe_text(package.get("name")),
            safe_text(package.get("typeLabel")),
            package.get("drawingCount") or len(package.get("items") or []),
            safe_text(package.get("statusLabel")),
            safe_text(package.get("createdByName")),
            format_timestamp(package.get("createdAt")),
            safe_text(package.get("ackSummary"), "—"),
            "；".join(safe_text(item) for item in warnings),
        ]
        write_row(sheet, row_index, row)
        if warnings:
            for col_index in range(1, len(headers) + 1):
                sheet.cell(row_index, col_index).fill = WARNING_FILL
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:I{max(2, len(packages) + 1)}"
    fit_columns(sheet, [14, 28, 12, 10, 12, 16, 20, 12, 58])


def write_summary_sheet(sheet, register):
    sheet["A1"] = "图纸目录导出摘要"
    sheet["A1"].font = Font(bold=True, size=15, color="172033")
    sheet["A1"].fill = TITLE_FILL
    sheet.merge_cells("A1:D1")
    completeness = register.get("completeness") if isinstance(register.get("completeness"), dict) else {}
    rows = [
        ("项目 ID", safe_text(register.get("projectId"))),
        ("导出时间", format_timestamp(register.get("generatedAt")) or datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("图纸总数", completeness.get("total", 0)),
        ("PDF + DWG 齐全", completeness.get("both", 0)),
        ("仅有 PDF", completeness.get("onlyPdf", 0)),
        ("仅有 DWG", completeness.get("onlyDwg", 0)),
        ("格式不一致", completeness.get("inconsistent", 0)),
    ]
    for row_index, (label, value) in enumerate(rows, start=3):
        sheet.cell(row_index, 1, label).font = Font(bold=True)
        sheet.cell(row_index, 1).fill = HEADER_FILL
        sheet.cell(row_index, 2, value)
        sheet.cell(row_index, 2).alignment = TEXT_WRAP
    fit_columns(sheet, [22, 42, 18, 18])


def write_header(sheet, row_index, headers):
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row_index, col_index, header)
        cell.font = Font(bold=True, color="172033")
        cell.fill = HEADER_FILL
        cell.alignment = TEXT_WRAP


def write_row(sheet, row_index, values):
    for col_index, value in enumerate(values, start=1):
        cell = sheet.cell(row_index, col_index, value)
        cell.alignment = TEXT_WRAP


def fit_columns(sheet, widths):
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


if __name__ == "__main__":
    main()
