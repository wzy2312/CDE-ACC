import json
import os
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from report_i18n import localize_workbook


HEADER_FILL = PatternFill("solid", fgColor="E8EEF7")
TITLE_FILL = PatternFill("solid", fgColor="DDE8F5")
SUCCESS_FILL = PatternFill("solid", fgColor="E8F6EF")
TEXT_WRAP = Alignment(vertical="top", wrap_text=True)


def main():
    if len(sys.argv) != 4:
        raise SystemExit("usage: export_quantity_takeoff.py <store_json> <task_id> <output_xlsx>")

    store_path, task_id, output_xlsx = sys.argv[1:]
    with open(store_path, "r", encoding="utf-8") as handle:
        store = json.load(handle)

    task = find_by_id(store.get("quantityTakeoffTasks", []), task_id)
    if not task:
        raise SystemExit("quantity takeoff task not found in store")

    document = find_by_id(store.get("documents", []), task.get("documentId")) or {}
    snapshots = [item for item in store.get("quantityPropertySnapshots", []) if item.get("taskId") == task_id]
    summaries = [item for item in store.get("quantitySummaries", []) if item.get("taskId") == task_id]

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "工程量汇总"
    write_summary_sheet(summary_sheet, document, task, summaries)

    detail_sheet = workbook.create_sheet("构件明细")
    write_detail_sheet(detail_sheet, document, task, snapshots)

    export_sheet = workbook.create_sheet("导出摘要")
    write_export_sheet(export_sheet, document, task, snapshots, summaries)

    os.makedirs(os.path.dirname(output_xlsx), exist_ok=True)
    localize_workbook(workbook)
    workbook.save(output_xlsx)


def safe_text(value, fallback=""):
    if value is None:
        return fallback
    text = str(value).strip()
    return text if text else fallback


def find_by_id(items, item_id):
    return next((item for item in items if safe_text(item.get("id")) == safe_text(item_id)), None)


def format_timestamp(value):
    raw = safe_text(value)
    if not raw:
        return ""
    try:
        dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return raw


def template_fields(task, snapshots):
    fields = (task.get("config") or {}).get("fields")
    if isinstance(fields, list) and fields:
        return [safe_text(item) for item in fields if safe_text(item)]
    keys = []
    seen = set()
    for snapshot in snapshots:
        for key in (snapshot.get("properties") or {}).keys():
            normalized = safe_text(key).lower()
            if normalized and normalized not in seen:
                seen.add(normalized)
                keys.append(safe_text(key))
    return keys[:40]


def write_meta_block(sheet, document, task):
    sheet["A1"] = "海水淡化工程量统计表"
    sheet["A1"].font = Font(bold=True, size=16, color="172033")
    sheet["A1"].fill = TITLE_FILL
    sheet.merge_cells("A1:I1")

    rows = [
        ("模型名称", safe_text(document.get("name"), "未命名模型"), "模型版本", safe_text(task.get("version"), safe_text(document.get("version"), "V1"))),
        ("模型 URN", safe_text(task.get("modelUrn")), "提取时间", format_timestamp(task.get("completedAt") or task.get("updatedAt"))),
        ("任务状态", safe_text(task.get("status")), "模板", safe_text((task.get("config") or {}).get("name"), "海水淡化工程量统计")),
    ]
    for index, row in enumerate(rows, start=2):
        for col, value in enumerate(row, start=1):
            cell = sheet.cell(index, col, value)
            cell.alignment = TEXT_WRAP
            if col in (1, 3):
                cell.font = Font(bold=True)
                cell.fill = HEADER_FILL


def write_summary_sheet(sheet, document, task, summaries):
    write_meta_block(sheet, document, task)
    metrics = metric_columns(summaries)
    headers = ["对象类型", "系统/区域", "专业", "材质/规格", "数量", "总面积(m²)", "总长度(m)", "总体积(m³)", *[metric["header"] for metric in metrics], "dbId 列表"]
    start_row = 6
    write_header(sheet, start_row, headers)

    for row_index, summary in enumerate(sorted_summaries(summaries), start=start_row + 1):
        values = [
            summary.get("elementType"),
            summary.get("floor"),
            summary.get("discipline"),
            summary.get("material"),
            summary.get("count"),
            summary.get("area"),
            summary.get("length"),
            summary.get("volume"),
            *[metric_value(summary, metric["key"]) for metric in metrics],
            ", ".join(str(item) for item in summary.get("dbIds", [])),
        ]
        write_row(sheet, row_index, values)

    apply_number_formats(sheet, start_row + 1, start_row + max(1, len(summaries)), list(range(5, 9 + len(metrics))))
    sheet.freeze_panes = "A7"
    fit_columns(sheet, [18, 16, 16, 18, 10, 12, 12, 12, *[14 for _ in metrics], 36])
    end_col = get_column_letter(len(headers))
    sheet.auto_filter.ref = f"A{start_row}:{end_col}{start_row + max(1, len(summaries))}"


def write_detail_sheet(sheet, document, task, snapshots):
    sheet["A1"] = "对象属性快照"
    sheet["A1"].font = Font(bold=True, size=15, color="172033")
    sheet["A1"].fill = TITLE_FILL
    sheet.merge_cells("A1:I1")
    sheet["A2"] = f"{safe_text(document.get('name'), '未命名模型')} · {safe_text(task.get('version'), 'V1')}"

    fields = template_fields(task, snapshots)
    headers = ["dbId", "对象名称", "对象类型", "系统/区域", "专业", "材质/规格", "面积(m²)", "长度(m)", "体积(m³)", "管径(mm)", "重量(kg)", *fields]
    start_row = 4
    write_header(sheet, start_row, headers)

    for row_index, snapshot in enumerate(sorted_snapshots(snapshots), start=start_row + 1):
        properties = snapshot.get("properties") or {}
        values = [
            snapshot.get("dbId"),
            snapshot.get("name"),
            snapshot.get("elementType"),
            snapshot.get("floor"),
            snapshot.get("discipline"),
            snapshot.get("material"),
            snapshot.get("area"),
            snapshot.get("length"),
            snapshot.get("volume"),
            snapshot.get("diameter"),
            snapshot.get("weight"),
            *[properties.get(field, "") for field in fields],
        ]
        write_row(sheet, row_index, values)

    apply_number_formats(sheet, start_row + 1, start_row + max(1, len(snapshots)), [1, 7, 8, 9, 10, 11])
    sheet.freeze_panes = "A5"
    widths = [10, 24, 18, 16, 16, 18, 12, 12, 12, 12, 12] + [18 for _ in fields]
    fit_columns(sheet, widths)
    end_col = get_column_letter(len(headers))
    sheet.auto_filter.ref = f"A{start_row}:{end_col}{start_row + max(1, len(snapshots))}"


def write_export_sheet(sheet, document, task, snapshots, summaries):
    sheet["A1"] = "导出摘要"
    sheet["A1"].font = Font(bold=True, size=15, color="172033")
    sheet["A1"].fill = TITLE_FILL
    sheet.merge_cells("A1:D1")
    rows = [
        ("模型名称", safe_text(document.get("name"), "未命名模型")),
        ("模型版本", safe_text(task.get("version"), safe_text(document.get("version"), "V1"))),
        ("模型 URN", safe_text(task.get("modelUrn"))),
        ("任务 ID", safe_text(task.get("id"))),
        ("任务状态", safe_text(task.get("status"))),
        ("提取完成时间", format_timestamp(task.get("completedAt") or task.get("updatedAt"))),
        ("对象快照数量", len(snapshots)),
        ("汇总行数量", len(summaries)),
        ("模板名称", safe_text((task.get("config") or {}).get("name"), "海水淡化工程量统计")),
        ("分组维度", " / ".join((task.get("config") or {}).get("groupBy") or [])),
        ("模板指标", " / ".join(metric["header"] for metric in metric_columns(summaries))),
        ("导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for index, (label, value) in enumerate(rows, start=3):
        sheet.cell(index, 1, label).font = Font(bold=True)
        sheet.cell(index, 1).fill = HEADER_FILL
        sheet.cell(index, 2, value)
        sheet.cell(index, 2).alignment = TEXT_WRAP
    fit_columns(sheet, [18, 64, 18, 18])


def sorted_summaries(summaries):
    return sorted(
        summaries,
        key=lambda item: (
            safe_text(item.get("elementType")),
            safe_text(item.get("floor")),
            safe_text(item.get("discipline")),
            safe_text(item.get("material")),
        ),
    )


def sorted_snapshots(snapshots):
    return sorted(
        snapshots,
        key=lambda item: (
            safe_text(item.get("elementType")),
            safe_text(item.get("floor")),
            int(item.get("dbId") or 0),
        ),
    )


def metric_columns(summaries):
    columns = []
    seen = set()
    for summary in summaries or []:
        metrics = summary.get("metrics") or {}
        for key, metric in metrics.items():
            normalized = safe_text(key).lower()
            if not normalized or normalized in seen:
                continue
            seen.add(normalized)
            label = safe_text(metric.get("label"), key)
            unit = safe_text(metric.get("unit"))
            columns.append({
                "key": key,
                "header": f"{label}({unit})" if unit else label,
            })
    return columns


def metric_value(summary, key):
    metric = (summary.get("metrics") or {}).get(key) or {}
    return metric.get("value", 0)


def write_header(sheet, row_index, headers):
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row_index, col_index, header)
        cell.font = Font(bold=True, color="172033")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_row(sheet, row_index, values):
    for col_index, value in enumerate(values, start=1):
        if isinstance(value, (dict, list)):
            value = json.dumps(value, ensure_ascii=False)
        cell = sheet.cell(row_index, col_index, value)
        cell.alignment = TEXT_WRAP


def apply_number_formats(sheet, start_row, end_row, columns):
    for row in range(start_row, end_row + 1):
        for column in columns:
            sheet.cell(row, column).number_format = "0.00"


def fit_columns(sheet, widths):
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


if __name__ == "__main__":
    main()
