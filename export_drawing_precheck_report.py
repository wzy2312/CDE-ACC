import json
import os
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from report_i18n import localize_workbook


HEADER_FILL = PatternFill("solid", fgColor="E8EEF7")
TITLE_FILL = PatternFill("solid", fgColor="DDE8F5")
ERROR_FILL = PatternFill("solid", fgColor="FCE8E8")
WARNING_FILL = PatternFill("solid", fgColor="FFF4D8")
INFO_FILL = PatternFill("solid", fgColor="EAF3FF")
TEXT_WRAP = Alignment(vertical="top", wrap_text=True)


def main():
    if len(sys.argv) != 4:
        raise SystemExit("usage: export_drawing_precheck_report.py <store_json> <task_id> <output_xlsx>")

    store_path, task_id, output_xlsx = sys.argv[1:]
    with open(store_path, "r", encoding="utf-8") as handle:
        store = json.load(handle)

    task = find_by_id(store.get("drawingPrecheckTasks", []), task_id)
    if not task:
        raise SystemExit("drawing precheck task not found in store")

    document = find_by_id(store.get("documents", []), task.get("fileId")) or {}
    results = [item for item in store.get("drawingPrecheckResults", []) if safe_text(item.get("taskId")) == safe_text(task_id)]

    workbook = Workbook()
    issue_sheet = workbook.active
    issue_sheet.title = "预审问题"
    write_issue_sheet(issue_sheet, document, task, results)

    summary_sheet = workbook.create_sheet("导出摘要")
    write_summary_sheet(summary_sheet, document, task, results)

    os.makedirs(os.path.dirname(output_xlsx), exist_ok=True)
    localize_workbook(workbook)
    workbook.save(output_xlsx)


def safe_text(value, fallback=""):
    if value is None:
        return fallback
    text = str(value).strip()
    return text if text else fallback


def find_by_id(items, item_id):
    return next((item for item in items if safe_text(item.get("id")) == safe_text(item_id)), None)


def format_timestamp(value):
    raw = safe_text(value)
    if not raw:
        return ""
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00")).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return raw


def level_label(value):
    return {
        "error": "阻断",
        "warning": "警告",
        "info": "提示",
    }.get(safe_text(value), safe_text(value, "提示"))


def check_type_label(value):
    return {
        "rule": "规则检查",
        "ai": "AI 审查",
        "visual": "视觉提示",
    }.get(safe_text(value), safe_text(value, "规则检查"))


def status_label(value):
    return {
        "open": "未处理",
        "resolved": "已修复",
        "accepted": "已接受",
        "false_positive": "误报",
    }.get(safe_text(value), safe_text(value, "未处理"))


def issue_link_label(result):
    annotation_id = safe_text(result.get("annotationId"))
    return f"已生成批注 {annotation_id}" if annotation_id else "仅报告项"


def position_label(result):
    position = result.get("position") if isinstance(result.get("position"), dict) else {}
    if not position:
        return ""
    x = float(position.get("x", 0) or 0) * 100
    y = float(position.get("y", 0) or 0) * 100
    width = float(position.get("width", 0) or 0) * 100
    height = float(position.get("height", 0) or 0) * 100
    return f"x={x:.1f}%, y={y:.1f}%, w={width:.1f}%, h={height:.1f}%"


def write_issue_sheet(sheet, document, task, results):
    headers = ["文件名称", "版本", "检查来源", "级别", "分类", "问题标题", "问题描述", "页码", "定位", "处理状态", "Issue 联动", "更新时间"]
    start_row = 1
    write_header(sheet, start_row, headers)

    for row_index, result in enumerate(sorted_results(results), start=start_row + 1):
        row = [
            safe_text(result.get("fileName"), safe_text(document.get("name"), safe_text(task.get("fileName"), ""))),
            safe_text(result.get("version"), safe_text(task.get("version"), "V1")),
            check_type_label(result.get("checkType")),
            level_label(result.get("level")),
            safe_text(result.get("category"), "general"),
            safe_text(result.get("title"), "未命名问题"),
            safe_text(result.get("description")),
            result.get("page") or 1,
            position_label(result),
            status_label(result.get("status")),
            issue_link_label(result),
            format_timestamp(result.get("updatedAt") or result.get("createdAt")),
        ]
        write_row(sheet, row_index, row)
        level = safe_text(result.get("level"))
        fill = ERROR_FILL if level == "error" else WARNING_FILL if level == "warning" else INFO_FILL
        for col_index in range(1, len(headers) + 1):
            sheet.cell(row_index, col_index).fill = fill

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A{start_row}:L{start_row + max(1, len(results))}"
    fit_columns(sheet, [24, 10, 14, 10, 16, 28, 46, 8, 28, 12, 22, 20])


def write_summary_sheet(sheet, document, task, results):
    sheet["A1"] = "导出摘要"
    sheet["A1"].font = Font(bold=True, size=15, color="172033")
    sheet["A1"].fill = TITLE_FILL
    sheet.merge_cells("A1:D1")

    counts = {
        "total": len(results),
        "error": sum(1 for item in results if safe_text(item.get("level")) == "error"),
        "warning": sum(1 for item in results if safe_text(item.get("level")) == "warning"),
        "open": sum(1 for item in results if safe_text(item.get("status")) == "open"),
        "linked": sum(1 for item in results if safe_text(item.get("annotationId"))),
    }
    rows = [
        ("文件名称", safe_text(document.get("name"), safe_text(task.get("fileName"), "未命名图纸"))),
        ("模型/图纸版本", safe_text(task.get("version"), safe_text(document.get("version"), "V1"))),
        ("任务 ID", safe_text(task.get("id"))),
        ("任务状态", safe_text(task.get("status"))),
        ("完成时间", format_timestamp(task.get("completedAt") or task.get("updatedAt"))),
        ("问题总数", counts["total"]),
        ("阻断项", counts["error"]),
        ("警告项", counts["warning"]),
        ("未处理项", counts["open"]),
        ("已生成 Issue 批注", counts["linked"]),
        ("导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for row_index, (label, value) in enumerate(rows, start=3):
        sheet.cell(row_index, 1, label).font = Font(bold=True)
        sheet.cell(row_index, 1).fill = HEADER_FILL
        sheet.cell(row_index, 2, value)
        sheet.cell(row_index, 2).alignment = TEXT_WRAP
    fit_columns(sheet, [20, 64, 18, 18])


def sorted_results(results):
    order = {"error": 0, "warning": 1, "info": 2}
    return sorted(
        results,
        key=lambda item: (
            order.get(safe_text(item.get("level")), 3),
            int(item.get("page") or 1),
            safe_text(item.get("category")),
            safe_text(item.get("title")),
        ),
    )


def write_header(sheet, row_index, headers):
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row_index, col_index, header)
        cell.font = Font(bold=True, color="172033")
        cell.fill = HEADER_FILL
        cell.alignment = TEXT_WRAP


def write_row(sheet, row_index, values):
    for col_index, value in enumerate(values, start=1):
        cell = sheet.cell(row_index, col_index, value)
        cell.alignment = TEXT_WRAP


def fit_columns(sheet, widths):
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


if __name__ == "__main__":
    main()
