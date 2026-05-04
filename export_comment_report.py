import json
import os
import posixpath
import re
import sys
import zipfile
from collections import Counter
from datetime import datetime
from xml.etree import ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


WORD_NS = {
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
}
SHEET_REL_NS = {
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}
PPT_NS = {
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "p": "http://schemas.openxmlformats.org/presentationml/2006/main",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def main():
    if len(sys.argv) != 5:
        raise SystemExit("usage: export_comment_report.py <store_json> <doc_id> <input_file> <output_xlsx>")

    store_path, doc_id, input_file, output_xlsx = sys.argv[1:]
    with open(store_path, "r", encoding="utf-8") as handle:
        store = json.load(handle)

    documents = store if isinstance(store, list) else store.get("documents", [])
    workflows = [] if isinstance(store, list) else store.get("workflows", [])

    document = next((item for item in documents if item.get("id") == doc_id), None)
    if not document:
        raise SystemExit("document not found in store")

    version_entry = current_version_entry(document)
    version_id = safe_text(document.get("currentVersionId"))
    ext = os.path.splitext(version_entry.get("name") or document.get("name") or input_file)[1].lower()
    native_rows, native_support = collect_native_comment_rows(input_file, ext, document, version_entry)

    rows = []
    rows.extend(collect_remark_rows(document, version_entry))
    rows.extend(collect_annotation_rows(document, version_id, version_entry))
    rows.extend(collect_workflow_rows(document, workflows, version_entry))
    rows.extend(native_rows)
    rows.sort(key=sort_key_for_row)

    workbook = Workbook()
    comment_sheet = workbook.active
    comment_sheet.title = "评论清单"
    write_comment_sheet(comment_sheet, rows)

    summary_sheet = workbook.create_sheet("导出摘要")
    write_summary_sheet(summary_sheet, document, version_entry, rows, native_support)

    workbook.save(output_xlsx)


def safe_text(value, fallback=""):
    if value is None:
        return fallback
    text = str(value).strip()
    return text if text else fallback


def normalize_space(value):
    return re.sub(r"\s+", " ", safe_text(value)).strip()


def multiline_text(value):
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    lines = [line.strip() for line in text.split("\n")]
    return "\n".join([line for line in lines if line])


def truncate(value, limit=180):
    text = normalize_space(value)
    if len(text) <= limit:
        return text
    return f"{text[:limit - 1]}…"


def format_timestamp(value):
    raw = safe_text(value)
    if not raw:
        return ""
    try:
        normalized = raw.replace("Z", "+00:00")
        dt = datetime.fromisoformat(normalized)
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return raw


def current_version_entry(document):
    current_version_id = safe_text(document.get("currentVersionId"))
    history = document.get("versionHistory") or []
    for entry in history:
        if safe_text(entry.get("id")) == current_version_id:
            return entry
    return {
        "id": current_version_id,
        "version": safe_text(document.get("version"), "V1"),
        "name": safe_text(document.get("name"), "未命名文件"),
    }


def annotation_status_label(value):
    return {
        "open": "未解决",
        "in_progress": "处理中",
        "resolved": "已解决",
    }.get(safe_text(value), safe_text(value))


def attachment_count(items):
    return len(items) if isinstance(items, list) else 0


def format_annotation_anchor(annotation):
    page = annotation.get("page")
    x = float(annotation.get("x", 0) or 0) * 100
    y = float(annotation.get("y", 0) or 0) * 100
    width = float(annotation.get("width", 0) or 0) * 100
    height = float(annotation.get("height", 0) or 0) * 100
    page_label = f"第 {int(page)} 页" if page else ""
    geometry = f"x={x:.1f}%, y={y:.1f}%, w={width:.1f}%, h={height:.1f}%"
    return " · ".join([item for item in [page_label, geometry] if item])


def file_kind_label(ext):
    if ext == ".pdf":
        return "PDF"
    if ext in {".doc", ".docx"}:
        return "Word"
    if ext in {".xls", ".xlsx", ".csv", ".ods"}:
        return "Excel"
    if ext in {".ppt", ".pptx", ".odp"}:
        return "PowerPoint"
    return ext.replace(".", "").upper() or "文件"


def base_row(document, version_entry, record_type, source, **extra):
    return {
        "文件名称": safe_text(document.get("name"), safe_text(version_entry.get("name"), "未命名文件")),
        "当前版本": safe_text(version_entry.get("version"), safe_text(document.get("version"), "V1")),
        "文件类型": file_kind_label(os.path.splitext(safe_text(document.get("name"), safe_text(version_entry.get("name"))))[1].lower()),
        "记录类型": record_type,
        "评论来源": source,
        "记录ID": safe_text(extra.get("record_id")),
        "状态": safe_text(extra.get("status")),
        "作者": safe_text(extra.get("actor")),
        "时间": format_timestamp(extra.get("created_at")),
        "页码/页签": safe_text(extra.get("page_or_sheet")),
        "锚点/定位": safe_text(extra.get("anchor")),
        "标题": safe_text(extra.get("title")),
        "内容": safe_text(extra.get("content")),
        "回复至": safe_text(extra.get("reply_to")),
        "附件数": int(extra.get("attachment_count") or 0),
        "版本ID": safe_text(extra.get("version_id"), safe_text(version_entry.get("id"))),
        "_sort_time": format_timestamp(extra.get("created_at")),
    }


def collect_remark_rows(document, version_entry):
    remark = multiline_text(document.get("remarks"))
    if not remark:
        return []
    return [
        base_row(
            document,
            version_entry,
            "文档备注",
            "系统备注",
            record_id=f"{safe_text(document.get('id'))}-remark",
            status="已记录",
            actor=safe_text(document.get("uploader"), safe_text(document.get("owner"), "系统")),
            created_at=document.get("updatedAt") or document.get("uploadedAt"),
            title="整份文件备注",
            content=remark,
        )
    ]


def collect_annotation_rows(document, current_version_id, version_entry):
    rows = []
    annotations = document.get("annotations") or []
    for annotation in annotations:
        annotation_version_id = safe_text(annotation.get("versionId"))
        if current_version_id and annotation_version_id and annotation_version_id != current_version_id:
            continue
        rows.append(
            base_row(
                document,
                version_entry,
                "系统批注",
                "PDF审阅" if file_kind_label(os.path.splitext(safe_text(document.get("name")))[1].lower()) == "PDF" else "系统审阅",
                record_id=annotation.get("id"),
                status=annotation_status_label(annotation.get("status") or ("resolved" if annotation.get("resolved") else "open")),
                actor=annotation.get("actor"),
                created_at=annotation.get("createdAt"),
                page_or_sheet=f"第 {int(annotation.get('page') or 1)} 页",
                anchor=format_annotation_anchor(annotation),
                title=safe_text(annotation.get("title"), "未命名批注"),
                content=multiline_text(annotation.get("note")),
                attachment_count=attachment_count(annotation.get("attachments")),
                version_id=annotation_version_id or current_version_id,
            )
        )
        for reply in annotation.get("replies") or []:
            rows.append(
                base_row(
                    document,
                    version_entry,
                    "批注回复",
                    "PDF审阅回复" if file_kind_label(os.path.splitext(safe_text(document.get("name")))[1].lower()) == "PDF" else "系统批注回复",
                    record_id=reply.get("id"),
                    status=annotation_status_label(annotation.get("status") or ("resolved" if annotation.get("resolved") else "open")),
                    actor=reply.get("actor"),
                    created_at=reply.get("createdAt"),
                    page_or_sheet=f"第 {int(annotation.get('page') or 1)} 页",
                    anchor=format_annotation_anchor(annotation),
                    title=safe_text(annotation.get("title"), "未命名批注"),
                    content=multiline_text(reply.get("content")),
                    reply_to=safe_text(annotation.get("title"), "未命名批注"),
                    attachment_count=attachment_count(reply.get("attachments")),
                    version_id=annotation_version_id or current_version_id,
                )
            )
    return rows


def collect_workflow_rows(document, workflows, version_entry):
    rows = []
    document_id = safe_text(document.get("id"))
    workflow_ids = {safe_text(item) for item in (document.get("workflowIds") or []) if safe_text(item)}
    for workflow in workflows or []:
        workflow_id = safe_text(workflow.get("id"))
        file_refs = workflow.get("fileRefs") or []
        linked = workflow_id in workflow_ids or any(safe_text(item.get("docId")) == document_id for item in file_refs)
        if not linked:
            continue
        for step in workflow.get("steps") or []:
            step_name = safe_text(step.get("name"), "审批节点")
            for reviewer in step.get("reviewers") or []:
                comment = multiline_text(reviewer.get("comment"))
                if not comment:
                    continue
                rows.append(
                    base_row(
                        document,
                        version_entry,
                        "流程审批意见",
                        "审批流程",
                        record_id=reviewer.get("id"),
                        status=safe_text(reviewer.get("status")),
                        actor=reviewer.get("name"),
                        created_at=reviewer.get("actedAt"),
                        page_or_sheet=step_name,
                        anchor=safe_text(workflow.get("workflowName"), safe_text(workflow.get("templateName"))),
                        title=" / ".join(
                            [
                                part
                                for part in [
                                    safe_text(workflow.get("workflowName"), safe_text(workflow.get("templateName"))),
                                    step_name,
                                ]
                                if part
                            ]
                        ),
                        content=comment,
                        version_id=safe_text(version_entry.get("id")),
                    )
                )
    return rows


def collect_native_comment_rows(input_file, ext, document, version_entry):
    if ext == ".docx":
        return extract_docx_rows(input_file, document, version_entry), "已提取 DOCX 原生评论"
    if ext == ".xlsx":
        return extract_xlsx_rows(input_file, document, version_entry), "已提取 XLSX 原生评论"
    if ext == ".pptx":
        return extract_pptx_rows(input_file, document, version_entry), "已提取 PPTX 原生评论"
    if ext in {".doc", ".xls", ".ppt", ".csv", ".ods", ".odp"}:
        return [], "当前格式暂不支持提取原生评论，报表仅包含系统批注与备注"
    return [], "当前文件无可提取的 Office 原生评论"


def extract_docx_rows(input_file, document, version_entry):
    if not zipfile.is_zipfile(input_file):
        return []
    with zipfile.ZipFile(input_file) as archive:
        if "word/comments.xml" not in archive.namelist():
            return []
        anchors = extract_docx_comment_anchors(archive)
        root = ET.fromstring(archive.read("word/comments.xml"))
        rows = []
        for comment in root.findall("w:comment", WORD_NS):
            comment_id = safe_text(comment.attrib.get(f"{{{WORD_NS['w']}}}id"))
            author = safe_text(comment.attrib.get(f"{{{WORD_NS['w']}}}author"), "未知作者")
            created_at = safe_text(comment.attrib.get(f"{{{WORD_NS['w']}}}date"))
            paragraphs = []
            for paragraph in comment.findall(".//w:p", WORD_NS):
                text = "".join((node.text or "") for node in paragraph.findall(".//w:t", WORD_NS))
                if normalize_space(text):
                    paragraphs.append(multiline_text(text))
            content = "\n".join(paragraphs)
            parent_id = ""
            for attr_key, attr_value in comment.attrib.items():
                if attr_key.endswith("parentId"):
                    parent_id = safe_text(attr_value)
                    break
            anchor = anchors.get(comment_id, {})
            rows.append(
                base_row(
                    document,
                    version_entry,
                    "Office原生评论回复" if parent_id else "Office原生评论",
                    "OnlyOffice原生评论",
                    record_id=comment_id,
                    status="已保存",
                    actor=author,
                    created_at=created_at,
                    page_or_sheet=safe_text(anchor.get("paragraph"), "正文"),
                    anchor=safe_text(anchor.get("anchor")),
                    title=f"Word 评论 {comment_id or '未命名'}",
                    content=content,
                    reply_to=parent_id,
                    version_id=safe_text(version_entry.get("id")),
                )
            )
        return rows


def extract_docx_comment_anchors(archive):
    anchors = {}
    if "word/document.xml" not in archive.namelist():
        return anchors
    root = ET.fromstring(archive.read("word/document.xml"))
    paragraphs = root.findall(".//w:body//w:p", WORD_NS)
    for index, paragraph in enumerate(paragraphs, start=1):
        text = "".join((node.text or "") for node in paragraph.findall(".//w:t", WORD_NS))
        paragraph_text = truncate(text or "（空段落）")
        for marker in paragraph.findall(".//w:commentRangeStart", WORD_NS):
            comment_id = safe_text(marker.attrib.get(f"{{{WORD_NS['w']}}}id"))
            if not comment_id or comment_id in anchors:
                continue
            anchors[comment_id] = {
                "paragraph": f"段落 {index}",
                "anchor": paragraph_text,
            }
    return anchors


def extract_xlsx_rows(input_file, document, version_entry):
    rows = []
    workbook = load_workbook(input_file, data_only=False)
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    comment = cell.comment
                    if not comment:
                        continue
                    rows.append(
                        base_row(
                            document,
                            version_entry,
                            "Office原生评论",
                            "OnlyOffice原生评论",
                            record_id=f"{sheet.title}!{cell.coordinate}",
                            status="已保存",
                            actor=safe_text(comment.author, "未知作者"),
                            created_at="",
                            page_or_sheet=sheet.title,
                            anchor=cell.coordinate,
                            title=f"{sheet.title}!{cell.coordinate}",
                            content=multiline_text(comment.text),
                            version_id=safe_text(version_entry.get("id")),
                        )
                    )
    finally:
        workbook.close()
    return rows


def extract_pptx_rows(input_file, document, version_entry):
    if not zipfile.is_zipfile(input_file):
        return []
    with zipfile.ZipFile(input_file) as archive:
        authors = extract_pptx_comment_authors(archive)
        rows = []
        for rel_name in archive.namelist():
            match = re.fullmatch(r"ppt/slides/_rels/slide(\d+)\.xml\.rels", rel_name)
            if not match:
                continue
            slide_number = int(match.group(1))
            relationships_root = ET.fromstring(archive.read(rel_name))
            comment_targets = []
            for relationship in relationships_root.findall("rel:Relationship", SHEET_REL_NS):
                rel_type = safe_text(relationship.attrib.get("Type"))
                if not rel_type.endswith("/comments"):
                    continue
                target = safe_text(relationship.attrib.get("Target"))
                if not target:
                    continue
                source_part = rel_name.replace("/_rels/", "/").replace(".rels", "")
                base_dir = posixpath.dirname(source_part)
                comment_path = posixpath.normpath(posixpath.join(base_dir, target))
                comment_targets.append(comment_path)
            if not comment_targets:
                continue
            slide_path = f"ppt/slides/slide{slide_number}.xml"
            slide_title = extract_pptx_slide_title(archive, slide_path)
            for comment_path in comment_targets:
                if comment_path not in archive.namelist():
                    continue
                comment_root = ET.fromstring(archive.read(comment_path))
                for comment in comment_root.findall(".//p:cm", PPT_NS):
                    author_id = safe_text(comment.attrib.get("authorId"))
                    created_at = safe_text(comment.attrib.get("dt"))
                    index = safe_text(comment.attrib.get("idx"))
                    text = safe_text(comment.findtext("p:text", default="", namespaces=PPT_NS))
                    if not text:
                        text = " ".join(
                            filter(
                                None,
                                [safe_text(node.text) for node in comment.findall(".//a:t", PPT_NS)],
                            )
                        )
                    position = comment.find("p:pos", PPT_NS)
                    anchor = slide_title
                    if position is not None:
                        coords = f"x={safe_text(position.attrib.get('x'))}, y={safe_text(position.attrib.get('y'))}"
                        anchor = f"{anchor} · {coords}" if anchor else coords
                    rows.append(
                        base_row(
                            document,
                            version_entry,
                            "Office原生评论",
                            "OnlyOffice原生评论",
                            record_id=f"slide-{slide_number}-{index or author_id}",
                            status="已保存",
                            actor=authors.get(author_id, "未知作者"),
                            created_at=created_at,
                            page_or_sheet=f"幻灯片 {slide_number}",
                            anchor=anchor,
                            title=f"幻灯片 {slide_number} 评论",
                            content=multiline_text(text),
                            version_id=safe_text(version_entry.get("id")),
                        )
                    )
        return rows


def extract_pptx_comment_authors(archive):
    if "ppt/commentAuthors.xml" not in archive.namelist():
        return {}
    root = ET.fromstring(archive.read("ppt/commentAuthors.xml"))
    authors = {}
    for author in root.findall(".//p:cmAuthor", PPT_NS):
        author_id = safe_text(author.attrib.get("id"))
        name = safe_text(author.attrib.get("name"), safe_text(author.attrib.get("initials"), "未知作者"))
        if author_id:
            authors[author_id] = name
    return authors


def extract_pptx_slide_title(archive, slide_path):
    if slide_path not in archive.namelist():
        return ""
    root = ET.fromstring(archive.read(slide_path))
    texts = [safe_text(node.text) for node in root.findall(".//a:t", PPT_NS) if safe_text(node.text)]
    return truncate(texts[0], 80) if texts else ""


def sort_key_for_row(row):
    return (
        safe_text(row.get("时间")),
        safe_text(row.get("评论来源")),
        safe_text(row.get("记录类型")),
        safe_text(row.get("记录ID")),
    )


def write_comment_sheet(sheet, rows):
    headers = [
        "文件名称",
        "当前版本",
        "文件类型",
        "记录类型",
        "评论来源",
        "记录ID",
        "状态",
        "作者",
        "时间",
        "页码/页签",
        "锚点/定位",
        "标题",
        "内容",
        "回复至",
        "附件数",
        "版本ID",
    ]
    sheet.append(headers)
    for header_cell in sheet[1]:
        header_cell.font = Font(bold=True, color="FFFFFF")
        header_cell.fill = PatternFill("solid", fgColor="1F5AA6")
        header_cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        sheet.append([row.get(header, "") for header in headers])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(2, sheet.max_row)}"
    sheet.sheet_view.showGridLines = True

    widths = {
        "A": 24,
        "B": 12,
        "C": 10,
        "D": 14,
        "E": 16,
        "F": 20,
        "G": 12,
        "H": 14,
        "I": 20,
        "J": 18,
        "K": 30,
        "L": 24,
        "M": 42,
        "N": 22,
        "O": 10,
        "P": 38,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_summary_sheet(sheet, document, version_entry, rows, native_support):
    counts_by_source = Counter(row.get("评论来源") for row in rows)
    counts_by_type = Counter(row.get("记录类型") for row in rows)
    meta_rows = [
        ("文件名称", safe_text(document.get("name"), "未命名文件")),
        ("当前版本", safe_text(version_entry.get("version"), safe_text(document.get("version"), "V1"))),
        ("导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("评论总数", str(len(rows))),
        ("原生评论提取", native_support),
        ("备注", "评论清单已汇总 PDF/系统批注、回复、备注、流程审批意见和 Office 原生评论。"),
        ("", ""),
        ("按来源统计", "数量"),
    ]
    for source, count in sorted(counts_by_source.items()):
        meta_rows.append((source, str(count)))
    meta_rows.append(("", ""))
    meta_rows.append(("按类型统计", "数量"))
    for record_type, count in sorted(counts_by_type.items()):
        meta_rows.append((record_type, str(count)))

    for index, (label, value) in enumerate(meta_rows, start=1):
        sheet.append([label, value])
        if label in {"按来源统计", "按类型统计"}:
            sheet[f"A{index}"].font = Font(bold=True, color="FFFFFF")
            sheet[f"B{index}"].font = Font(bold=True, color="FFFFFF")
            sheet[f"A{index}"].fill = PatternFill("solid", fgColor="1F5AA6")
            sheet[f"B{index}"].fill = PatternFill("solid", fgColor="1F5AA6")
        elif label:
            sheet[f"A{index}"].font = Font(bold=True)
        sheet[f"A{index}"].alignment = Alignment(vertical="top", wrap_text=True)
        sheet[f"B{index}"].alignment = Alignment(vertical="top", wrap_text=True)

    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 72


if __name__ == "__main__":
    main()
