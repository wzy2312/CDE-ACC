import json
import os
import sys
from datetime import datetime
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfgen import canvas

from pdf_font_utils import register_pdf_font
from report_i18n import localize_text, localize_workbook

FONT_NAME = "ClashHeatmapArialUnicode"
PAGE_WIDTH, PAGE_HEIGHT = A4
MARGIN_X = 46
TOP_Y = PAGE_HEIGHT - 54
LINE_HEIGHT = 17
HEADER_FILL = PatternFill("solid", fgColor="E8EEF7")
HOT_FILL = PatternFill("solid", fgColor="F8D7DA")
TEXT_WRAP = Alignment(vertical="top", wrap_text=True)


def main():
    if len(sys.argv) != 5:
        raise SystemExit("usage: export_model_clash_heatmap.py <store_json> <heatmap_id> <kind> <output>")

    store_path, heatmap_id, kind, output_path = sys.argv[1:]
    with open(store_path, "r", encoding="utf-8") as handle:
        store = json.load(handle)

    heatmap = find_by_id(store.get("modelClashHeatmaps", []), heatmap_id)
    if not heatmap:
        raise SystemExit("model clash heatmap not found in store")

    cells = [item for item in store.get("modelClashHeatmapCells", []) if safe_text(item.get("heatmapId")) == heatmap_id]
    hotspots = [item for item in store.get("modelClashHotspots", []) if safe_text(item.get("heatmapId")) == heatmap_id]
    records = [item for item in store.get("modelClashRecords", []) if safe_text(item.get("runId")) == safe_text(heatmap.get("clashTaskId"))]
    runs = store.get("modelClashRuns", [])
    run = find_by_id(runs, heatmap.get("clashTaskId")) or {}

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    if kind == "matrix":
        export_matrix(output_path, heatmap, run, records)
    elif kind == "hotspots":
        export_hotspots_pdf(output_path, heatmap, run, hotspots)
    elif kind == "snapshot":
        export_snapshot_svg(output_path, heatmap, cells, hotspots)
    else:
        raise SystemExit(f"unsupported export kind: {kind}")


def safe_text(value, fallback=""):
    if value is None:
        return fallback
    text = str(value).strip()
    return text if text else fallback


def find_by_id(items, item_id):
    return next((item for item in items if safe_text(item.get("id")) == safe_text(item_id)), None)


def format_datetime(value):
    raw = safe_text(value)
    if not raw:
        return ""
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00")).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return raw


def pair_key(left, right):
    return ":".join(sorted([safe_text(left, "general").lower(), safe_text(right, "general").lower()]))


def matrix_from_records(records):
    disciplines = sorted({
        safe_text(record.get("disciplineA"), "general").lower()
        for record in records
    } | {
        safe_text(record.get("disciplineB"), "general").lower()
        for record in records
    })
    matrix = {row: {col: None if row == col else 0 for col in disciplines} for row in disciplines}
    for record in records:
        left = safe_text(record.get("disciplineA"), "general").lower()
        right = safe_text(record.get("disciplineB"), "general").lower()
        if left and right and left != right:
            matrix[left][right] = (matrix[left][right] or 0) + 1
            matrix[right][left] = (matrix[right][left] or 0) + 1
    return disciplines, matrix


def export_matrix(output_path, heatmap, run, records):
    disciplines, matrix = matrix_from_records(records)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "专业碰撞矩阵"
    sheet["A1"] = "专业碰撞矩阵"
    sheet["A1"].font = Font(bold=True, size=16, color="172033")
    sheet["A2"] = f"检测任务：{safe_text(run.get('id'), safe_text(heatmap.get('clashTaskId')))}"
    sheet["A3"] = f"网格精度：{heatmap.get('gridSize', 1)}m · 导出时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    start_row = 5
    headers = ["专业", *disciplines]
    write_header(sheet, start_row, headers)
    max_count = max([matrix[row][col] or 0 for row in disciplines for col in disciplines] or [0])
    for row_index, row_name in enumerate(disciplines, start=start_row + 1):
        sheet.cell(row_index, 1, row_name)
        sheet.cell(row_index, 1).font = Font(bold=True)
        for col_index, col_name in enumerate(disciplines, start=2):
            value = matrix[row_name][col_name]
            cell = sheet.cell(row_index, col_index, "—" if value is None else value)
            if value and max_count:
                ratio = min(1, value / max_count)
                cell.fill = heat_fill(ratio)

    sheet.freeze_panes = "B6"
    for col in range(1, len(headers) + 1):
        sheet.column_dimensions[sheet.cell(1, col).column_letter].width = 16
    localize_workbook(workbook)
    workbook.save(output_path)


def heat_fill(ratio):
    if ratio >= 0.75:
        return PatternFill("solid", fgColor="D34A4A")
    if ratio >= 0.5:
        return PatternFill("solid", fgColor="F5B535")
    if ratio >= 0.25:
        return PatternFill("solid", fgColor="22A45D")
    return PatternFill("solid", fgColor="9CC9FF")


def write_header(sheet, row_index, headers):
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row_index, col_index, header)
        cell.font = Font(bold=True, color="172033")
        cell.fill = HEADER_FILL
        cell.alignment = TEXT_WRAP


def export_hotspots_pdf(output_path, heatmap, run, hotspots):
    font_name = register_font()
    pdf = canvas.Canvas(output_path, pagesize=A4)
    pdf.setTitle("Clash Heatmap Hotspot Report")
    cursor_y = TOP_Y
    pdf.setFont(font_name, 18)
    pdf.drawString(MARGIN_X, cursor_y, localize_text("碰撞热力图热点区域报告"))
    cursor_y -= 28
    pdf.setFont(font_name, 10.5)
    lines = [
        f"热力图 ID：{heatmap.get('id')}",
        f"检测任务：{safe_text(run.get('id'), safe_text(heatmap.get('clashTaskId')))}",
        f"网格精度：{heatmap.get('gridSize', 1)}m",
        f"生成时间：{format_datetime(heatmap.get('createdAt'))}",
        f"导出时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
    ]
    cursor_y = draw_lines(pdf, cursor_y, lines)
    cursor_y -= 8
    pdf.setFont(font_name, 13)
    pdf.drawString(MARGIN_X, cursor_y, localize_text("Top 热点区域"))
    cursor_y -= 20
    pdf.setFont(font_name, 10.5)
    for index, hotspot in enumerate(sorted(hotspots, key=lambda item: item.get("clashCount", 0), reverse=True), start=1):
        bbox = hotspot.get("bbox") or {}
        lines = [
            f"{index}. 碰撞数量：{hotspot.get('clashCount', 0)} · 未关闭 Issue：{hotspot.get('openIssueCount', 0)}",
            f"专业对：{', '.join(hotspot.get('disciplinePairs') or []) or '—'}",
            f"区域边界：min {bbox.get('min', [0, 0, 0])} / max {bbox.get('max', [0, 0, 0])}",
        ]
        cursor_y = draw_lines(pdf, cursor_y, lines)
        cursor_y -= 6
        if cursor_y < 80:
            pdf.showPage()
            cursor_y = TOP_Y
            pdf.setFont(font_name, 10.5)
    pdf.save()


def register_font():
    return register_pdf_font(FONT_NAME)


def draw_lines(pdf, cursor_y, lines):
    font_name = register_font()
    for line in lines:
        if cursor_y < 60:
            pdf.showPage()
            cursor_y = TOP_Y
            pdf.setFont(font_name, 10.5)
        pdf.drawString(MARGIN_X, cursor_y, localize_text(safe_text(line, "—"))[:110])
        cursor_y -= LINE_HEIGHT
    return cursor_y


def export_snapshot_svg(output_path, heatmap, cells, hotspots):
    width = 960
    height = 620
    plot_x = 70
    plot_y = 90
    plot_w = 680
    plot_h = 420
    if cells:
        min_x = min(float(cell.get("x", 0)) for cell in cells)
        max_x = max(float(cell.get("x", 0)) + float(cell.get("gridSize", heatmap.get("gridSize", 1))) for cell in cells)
        min_y = min(float(cell.get("y", 0)) for cell in cells)
        max_y = max(float(cell.get("y", 0)) + float(cell.get("gridSize", heatmap.get("gridSize", 1))) for cell in cells)
    else:
        min_x, max_x, min_y, max_y = 0, 1, 0, 1
    span_x = max(1, max_x - min_x)
    span_y = max(1, max_y - min_y)
    max_density = max([int(cell.get("density", 0) or 0) for cell in cells] or [1])
    rects = []
    for cell in cells:
        grid = float(cell.get("gridSize", heatmap.get("gridSize", 1)) or 1)
        x = plot_x + ((float(cell.get("x", 0)) - min_x) / span_x) * plot_w
        y = plot_y + plot_h - ((float(cell.get("y", 0)) - min_y + grid) / span_y) * plot_h
        w = max(6, grid / span_x * plot_w)
        h = max(6, grid / span_y * plot_h)
        density = int(cell.get("density", 0) or 0)
        opacity = 0.28 + 0.5 * min(1, density / max_density)
        color = escape(safe_text(cell.get("color"), "#237df0"))
        rects.append(f'<rect x="{x:.2f}" y="{y:.2f}" width="{w:.2f}" height="{h:.2f}" fill="{color}" opacity="{opacity:.2f}" stroke="#ffffff" stroke-width="1"/>')
    hotspot_labels = []
    for index, hotspot in enumerate(hotspots[:5], start=1):
        center = hotspot.get("center") or [0, 0, 0]
        x = plot_x + ((float(center[0]) - min_x) / span_x) * plot_w
        y = plot_y + plot_h - ((float(center[1]) - min_y) / span_y) * plot_h
        hotspot_labels.append(f'<circle cx="{x:.2f}" cy="{y:.2f}" r="13" fill="none" stroke="#172033" stroke-width="2"/><text x="{x - 4:.2f}" y="{y + 4:.2f}" font-size="12" font-weight="700">{index}</text>')

    svg = f'''<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">
  <rect width="{width}" height="{height}" fill="#f5f7fb"/>
  <text x="46" y="42" font-family="Arial, sans-serif" font-size="24" font-weight="700" fill="#172033">{escape(localize_text("碰撞热力图截图"))}</text>
  <text x="46" y="66" font-family="Arial, sans-serif" font-size="13" fill="#64748b">{escape(localize_text(f'热力图 {safe_text(heatmap.get("id"))} · 网格 {heatmap.get("gridSize", 1)}m · {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'))}</text>
  <rect x="{plot_x}" y="{plot_y}" width="{plot_w}" height="{plot_h}" fill="#ffffff" stroke="#cbd5e1"/>
  {''.join(rects)}
  {''.join(hotspot_labels)}
  <text x="{plot_x}" y="{plot_y + plot_h + 34}" font-family="Arial, sans-serif" font-size="12" fill="#64748b">{escape(localize_text(f'比例尺：单元格边长 {heatmap.get("gridSize", 1)}m，颜色由蓝、绿、黄到红表示碰撞密度升高。'))}</text>
  <g transform="translate(790 112)">
    <text x="0" y="-18" font-family="Arial, sans-serif" font-size="14" font-weight="700" fill="#172033">{escape(localize_text("图例"))}</text>
    <rect x="0" y="0" width="28" height="18" fill="#237df0" opacity="0.72"/><text x="40" y="14" font-family="Arial, sans-serif" font-size="12" fill="#64748b">{escape(localize_text("低密度"))}</text>
    <rect x="0" y="34" width="28" height="18" fill="#22a45d" opacity="0.72"/><text x="40" y="48" font-family="Arial, sans-serif" font-size="12" fill="#64748b">{escape(localize_text("中低"))}</text>
    <rect x="0" y="68" width="28" height="18" fill="#f5b535" opacity="0.72"/><text x="40" y="82" font-family="Arial, sans-serif" font-size="12" fill="#64748b">{escape(localize_text("中高"))}</text>
    <rect x="0" y="102" width="28" height="18" fill="#d34a4a" opacity="0.72"/><text x="40" y="116" font-family="Arial, sans-serif" font-size="12" fill="#64748b">{escape(localize_text("高密度"))}</text>
  </g>
</svg>'''
    with open(output_path, "w", encoding="utf-8") as handle:
        handle.write(svg)


if __name__ == "__main__":
    main()
